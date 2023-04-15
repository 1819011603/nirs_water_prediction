import pathlib

import numpy as np
from openpyxl import load_workbook
from sklearn.metrics import r2_score, mean_squared_error

from nirs.parameters import*
def getDataIndex(x0, index):
    if len(index) == len(x0[0]):
        return x0
    index = np.array(index,copy=False)
    return x0[:,index]


def get_log_name(pre="recode",suff= "log",dir_path="./",filter_method=None):
    import pathlib
    import re
    kk = re.compile("(\d+)")
    o=[]
    # print(suff)
    if suff.rfind(".") >=0 :
        o = [str(i.stem) for i in pathlib.Path(dir_path).glob("{}*{}".format(pre,suff[suff.rfind(".")+1:]))]
    else:
        o = [str(i.stem) for i in pathlib.Path(dir_path).glob("{}*{}".format(pre, "." + suff))]
    import datetime
    day = str(datetime.date.today())
    day = day[day.index('-')+1:].replace("-","_")
    max1 = 0
    for po in o:
        u = re.search(kk, po)
        if u != None:
            m = int(u.group(0))
            max1 = max(m, max1)
    f = pathlib.Path(dir_path)
    if not f.exists():
        f.mkdir(parents=True)
    if suff.rfind(".") < 0:

        return "{}/{}{}_{}.{}".format(f,pre,max1 + 1,day,suff)
    else:
        # return "{}/{}{}_{}{}".format(f,pre,max1 + 1,day,suff)
        filter_methods=[]
        if filter_method is None:
            pass
        return "{}/{}{}_{}{}".format(f,pre,max1 + 1,day,suff)

# def

def filter(x0,filter_method=None):
    x0 = np.array(x0,copy=True)
    for method in filter_method:
        x0 = method(x0)
    return x0

def getRR_RMSE_RPD(y_test,y_pred_test):
    r2_test = r2_score(y_test, y_pred_test)
    rmsep = np.sqrt(mean_squared_error(y_test, y_pred_test))
    RPD = np.std(y_test) / rmsep
    return r2_test,rmsep,RPD

def save2excel(row, header, path='tmp.xlsx'):
    if para.path != None:
        path = para.path

    path = "./result/" + path
    total = 3
    time1 = total

    complete = False
    while time1 > 0 and not  complete:
        try:
            create_excel(path)

            # 打开已有的Excel文件
            wb = load_workbook(filename=path)

            # 获取需要写入数据的sheet
            sheet = wb['Sheet']

            # 获取已有数据的最大行数
            max_row = sheet.max_row

            if max_row <= 1:
                sheet.append(header)
            sheet.append(row)

            wb.save(path)
            complete = True
            print(f"save in {path}")
        except Exception as e:
            if time1 == 1:
                raise str(e)
            else:
                print(f"{path} 第{total-time1+1}保存失败, 1s后重试 ")
                import time
                time.sleep(1)
        finally:
            time1-=1


def create_excel(path):
    import os
    from openpyxl import Workbook

    # 定义Excel文件路径
    file_path = path

    # 判断文件是否存在，不存在则创建
    if not os.path.exists(file_path):
        import pathlib
        file_dir = pathlib.Path(file_path).parent
        if not file_dir.exists():
            file_dir.mkdir(parents=True)



        # 创建一个新的Excel文件
        wb = Workbook()
        ws = wb.active
        # 保存Excel文件
        wb.save(os.path.abspath(file_path))
        print(f"{os.path.abspath(file_path)} created.")


def clear_excel(path):
    wb = load_workbook(filename=path)

    # 选择要清空的工作表
    ws = wb.active

    # 清空每个单元格的值
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    # 删除所有的行和列
    ws.delete_cols(1, ws.max_column)
    ws.delete_rows(1, ws.max_row)

    # 保存修改后的xlsx文件
    wb.save(filename=path)